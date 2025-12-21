"""
MongoDB-based event views - Hoàn toàn thay thế MySQL
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from applications.permissions import IsAdminOrTeacherOrDormSupervisor
from datetime import datetime, timedelta
import uuid
import logging

from applications.common.mongo import get_mongo_collection, to_plain
from applications.common.responses import ok, created, bad_request, not_found, server_error
from bson import ObjectId

logger = logging.getLogger(__name__)

# =============================================================================
# EVENT TYPES - MongoDB
# =============================================================================

@api_view(['GET'])
@permission_classes([])  # Public API
def mongo_event_types_list(request):
    """Lấy danh sách loại sự kiện từ MongoDB"""
    try:
        user = request.user
        user_role = getattr(user, 'role', 'student')
        
        coll = get_mongo_collection('event_types')
        
        # Filter theo role
        if user_role == 'student':
            # Học sinh chỉ thấy loại sự kiện được phép tạo
            query = {
                'allowed_roles': {'$in': ['student', 'both']},
                'is_active': True
            }
        elif user_role in ['teacher', 'admin']:
            # Giáo viên và admin thấy tất cả
            query = {'is_active': True}
        else:
            query = {'is_active': True}
        
        docs = list(coll.find(query).sort('name', 1))
        out = []
        for d in docs:
            doc = to_plain(d)
            # Sử dụng key có sẵn hoặc tạo từ name
            if 'key' not in doc:
                doc['key'] = doc.get('name', '').lower().replace(' ', '_').replace('/', '_')
            out.append(doc)
        
        return ok(out)
        
    except Exception as exc:
        logger.exception('mongo_event_types_list error')
        return server_error(exc)

@api_view(['GET'])
@permission_classes([])  # Public API
def mongo_event_types_template(request):
    """Trả nguyên vẹn 1 template (object) từ collection 'event_template'."""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        tmpl_coll = get_mongo_collection('event_template')
        doc = tmpl_coll.find_one({})
        if not doc:
            return not_found('Template is empty')
        return ok(to_plain(doc))
    except Exception as exc:
        logger.exception('mongo_event_types_template error')
        return server_error(exc)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mongo_event_types_detail(request, pk):
    """Chi tiết loại sự kiện từ MongoDB"""
    try:
        coll = get_mongo_collection('event_types')
        
        # Tìm theo ID (có thể là string hoặc ObjectId)
        try:
            doc = coll.find_one({'_id': ObjectId(pk)})
        except:
            doc = coll.find_one({'id': pk})
        
        if not doc:
            return not_found('Loại sự kiện không tồn tại')
        
        return ok(to_plain(doc))
        
    except Exception as exc:
        logger.exception('mongo_event_types_detail error')
        return server_error(exc)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def mongo_event_types_update(request, pk):
    """Cập nhật loại sự kiện trong MongoDB"""
    try:
        coll = get_mongo_collection('event_types')
        
        # Tìm theo ID
        try:
            doc = coll.find_one({'_id': ObjectId(pk)})
        except:
            doc = coll.find_one({'id': pk})
        
        if not doc:
            return Response({'detail': 'Loại sự kiện không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
        
        # Cập nhật
        update_data = request.data.copy()
        update_data['updated_at'] = datetime.now().isoformat()
        
        try:
            coll.update_one({'_id': ObjectId(pk)}, {'$set': update_data})
        except:
            coll.update_one({'id': pk}, {'$set': update_data})
        
        return ok({'message': 'Cập nhật loại sự kiện thành công'})
        
    except Exception as exc:
        logger.exception('mongo_event_types_update error')
        return server_error(exc)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def mongo_event_types_delete(request, pk):
    """Xóa loại sự kiện trong MongoDB"""
    try:
        coll = get_mongo_collection('event_types')
        
        # Tìm và xóa
        try:
            result = coll.delete_one({'_id': ObjectId(pk)})
        except:
            result = coll.delete_one({'id': pk})
        
        if result.deleted_count == 0:
            return not_found('Loại sự kiện không tồn tại')
        
        return ok({'message': 'Xóa loại sự kiện thành công'})
        
    except Exception as exc:
        logger.exception('mongo_event_types_delete error')
        return server_error(exc)

# =============================================================================
# EVENTS - MongoDB (Optimized for daily storage)
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mongo_events_optimized_list(request):
    """Lấy danh sách events tối ưu hóa cho 7 tiết học mỗi ngày"""
    try:
        user = request.user
        user_role = getattr(user, 'role', 'student')
        
        classroom_id = request.query_params.get('classroom_id')
        date = request.query_params.get('date')
        include_sudden = request.query_params.get('include_sudden', 'false').lower() == 'true'
        include_bonus = request.query_params.get('include_bonus', 'false').lower() == 'true'
        
        # Pagination parameters
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        page_size = min(page_size, 100)  # Limit max page size to 100
        
        coll = get_mongo_collection('events')
        query = {}
        
        # Filter theo role
        logger.info(f"User role: {user_role}, User ID: {user.id}")

        if classroom_id:
            query['classroom_id'] = classroom_id
        if date:
            query['date'] = date

        if user_role == 'teacher':
            # Teacher chỉ xem events của lớp mình chủ nhiệm
            classrooms_coll = get_mongo_collection('classrooms')
            teacher_classrooms = list(classrooms_coll.find({'homeroom_teacher_id': str(user.id)}))
            teacher_classroom_ids = [str(c['_id']) for c in teacher_classrooms]
            logger.info(f"Teacher classrooms found: {teacher_classroom_ids}")
            if teacher_classroom_ids:
                query['classroom_id'] = {'$in': teacher_classroom_ids}
            else:
                # Nếu teacher không có lớp chủ nhiệm, trả về empty
                logger.info("Teacher has no homeroom classes, returning empty")
                return Response({
                    'results': [],
                    'count': 0,
                    'page': page,
                    'page_size': page_size,
                    'total_pages': 0,
                    'next': None,
                    'previous': None
                }, status=status.HTTP_200_OK)
        elif user_role == 'student':
            # Student chỉ xem events của lớp mình
            # Lấy classroom_id trực tiếp từ user document trong collection users
            users_coll = get_mongo_collection('users')
            try:
                user_doc = users_coll.find_one({'_id': ObjectId(user.id)})
            except:
                user_doc = users_coll.find_one({'id': str(user.id)})
            
            if not user_doc:
                logger.info("Student user not found, returning empty")
                return Response({
                    'results': [],
                    'count': 0,
                    'page': page,
                    'page_size': page_size,
                    'total_pages': 0,
                    'next': None,
                    'previous': None
                }, status=status.HTTP_200_OK)
            
            student_classroom_id = user_doc.get('classroom_id')
            logger.info(f"Student classroom_id from user doc: {student_classroom_id}")
            
            if student_classroom_id:
                query['classroom_id'] = student_classroom_id
            else:
                # Nếu student không có lớp, trả về empty
                logger.info("Student has no classroom_id, returning empty")
                return Response({
                    'results': [],
                    'count': 0,
                    'page': page,
                    'page_size': page_size,
                    'total_pages': 0,
                    'next': None,
                    'previous': None
                }, status=status.HTTP_200_OK)
        # Admin xem tất cả (không filter)
        
        logger.info(f"Final query: {query}")

        # Get all documents matching query (không paginate ngay vì cần filter trước)
        # Để tối ưu, ta vẫn có thể paginate nhưng sẽ phải load nhiều hơn để đảm bảo có đủ documents sau filter
        # Tạm thời load nhiều hơn một chút để đảm bảo có đủ kết quả sau filter
        # Calculate skip
        skip = (page - 1) * page_size
        # Load thêm một số documents để đảm bảo sau khi filter vẫn có đủ page_size documents
        load_limit = page_size * 3  # Load gấp 3 lần để đảm bảo có đủ sau filter
        
        docs = list(coll.find(query).sort('date', -1).skip(skip).limit(load_limit))
        
        out = []
        for d in docs:
            t = to_plain(d)
            t['created_at'] = t.get('created_at') or ''
            t['updated_at'] = t.get('updated_at') or t['created_at']
            
            # Loại bỏ period "attendance", "violation_sudden", "bonus_sudden" khỏi hoạt động trong ngày
            # Trừ khi include_sudden=True (để hiển thị trang vi phạm đột xuất) hoặc include_bonus=True (để hiển thị trang điểm cộng)
            periods = t.get('periods', {})
            if isinstance(periods, dict):
                if include_sudden:
                    # Chỉ hiển thị periods["violation_sudden"], loại bỏ các periods khác
                    filtered_periods = {k: v for k, v in periods.items() if k == 'violation_sudden'}
                elif include_bonus:
                    # Chỉ hiển thị periods["bonus_sudden"], loại bỏ các periods khác
                    filtered_periods = {k: v for k, v in periods.items() if k == 'bonus_sudden'}
                else:
                    # Loại bỏ period "attendance", "violation_sudden", "bonus_sudden" (và "sudden" cho backward compatibility) khỏi hoạt động trong ngày
                    filtered_periods = {k: v for k, v in periods.items() if k not in ['attendance', 'violation_sudden', 'bonus_sudden', 'sudden']}
                
                # Chỉ thêm document vào kết quả nếu còn periods sau khi filter
                if filtered_periods:
                    t['periods'] = filtered_periods
                    out.append(t)
                    # Dừng khi đã có đủ page_size documents
                    if len(out) >= page_size:
                        break
        
        # Count tổng số documents có events sau khi filter
        # Để chính xác, ta cần đếm tất cả documents (không paginate)
        # Nhưng điều này tốn kém, nên ta sẽ estimate dựa trên tỷ lệ
        # Hoặc đơn giản hơn: chỉ trả về count của page hiện tại
        filtered_count = len(out)
        
        # Build pagination URLs
        base_url = request.build_absolute_uri().split('?')[0]
        params = request.GET.copy()
        
        # Estimate total pages dựa trên tỷ lệ filtered/total trong page này
        # Nếu có ít hơn page_size documents sau filter, có thể đã hết
        has_more = len(docs) == load_limit and filtered_count >= page_size
        
        next_url = None
        if has_more:
            params['page'] = page + 1
            next_url = f"{base_url}?{params.urlencode()}"
        
        previous_url = None
        if page > 1:
            params['page'] = page - 1
            previous_url = f"{base_url}?{params.urlencode()}"
        
        return Response({
            'results': out,
            'count': filtered_count,  # Count của page hiện tại sau filter
            'page': page,
            'page_size': page_size,
            'total_pages': 0 if filtered_count == 0 else (page + (1 if has_more else 0)),  # Estimate
            'next': next_url,
            'previous': previous_url
        }, status=status.HTTP_200_OK)
        
    except Exception as exc:
        logger.exception('mongo_events_optimized_list error')
        return server_error(exc)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mongo_events_optimized_detail(request):
    """Lấy chi tiết events theo ID hoặc date + classroom_id (trả về 1 object)."""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        from bson import ObjectId
        
        event_id = request.query_params.get('id')
        date = request.query_params.get('date')
        classroom_id = request.query_params.get('classroom_id')
        
        coll = get_mongo_collection('events')
        
        # Nếu có id, tìm theo id (ưu tiên)
        if event_id:
            try:
                d = coll.find_one({'_id': ObjectId(event_id)})
            except Exception:
                return bad_request('ID không hợp lệ')
        # Nếu không có id nhưng có date và classroom_id, tìm theo date và classroom_id
        elif date and classroom_id:
            try:
                d = coll.find_one({
                    'date': date,
                    'classroom_id': classroom_id
                })
            except Exception:
                return bad_request('Tham số không hợp lệ')
        else:
            return bad_request('Cần có id hoặc (date và classroom_id)')
        
        if not d:
            return not_found('Không tìm thấy dữ liệu')
        
        t = to_plain(d)
        
        # Enrich periods with student_name and event_type_name
        from bson import ObjectId
        students_coll = get_mongo_collection('students')
        event_types_coll = get_mongo_collection('event_types')
        
        # Collect all student IDs and event type keys from periods
        student_ids = set()
        event_type_keys = set()
        periods = t.get('periods', {})
        
        for period_events in periods.values():
            for event in period_events:
                if event.get('student_id'):
                    student_ids.add(event['student_id'])
                if event.get('event_type_key'):
                    event_type_keys.add(event['event_type_key'])
        
        # Batch lookup students
        student_map = {}
        if student_ids:
            object_ids = []
            for sid in student_ids:
                try:
                    object_ids.append(ObjectId(sid))
                except Exception:
                    pass
            
            for sd in students_coll.find({'_id': {'$in': object_ids}}):
                student_map[str(sd.get('_id'))] = (
                    sd.get('user', {}).get('full_name') or
                    (f"{sd.get('user', {}).get('first_name','')} {sd.get('user', {}).get('last_name','')}".strip())
                )
        
        # Batch lookup event types
        type_map = {}
        if event_type_keys:
            for et in event_types_coll.find({'key': {'$in': list(event_type_keys)}}):
                type_map[et.get('key')] = et.get('name')
        
        for period_num, period_events in periods.items():
            for event in period_events:
                # Add student_name
                if event.get('student_id') and event['student_id'] in student_map:
                    event['student_name'] = student_map[event['student_id']]
                
                # Add event_type_name
                event_type_key = event.get('event_type_key')
                if event_type_key:
                    if event_type_key in type_map:
                        event['event_type_name'] = type_map[event_type_key]
                    elif event_type_key == 'custom_bonus_point':
                        # Điểm cộng đột xuất: sử dụng description làm tên nếu không có trong event_types
                        event['event_type_name'] = event.get('description', 'Điểm cộng đột xuất')
        return ok(t)
    except Exception as exc:
        logging.getLogger(__name__).exception('mongo_events_optimized_detail error')
        return server_error(exc)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mongo_events_optimized_create(request):
    """Tạo events tối ưu hóa cho 7 tiết học mỗi ngày"""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        from bson import ObjectId
        from datetime import datetime
        
        user = request.user
        events_data = request.data.get('events', [])
        periods_payload = request.data.get('periods')  # optional grouped-by-period payload
        
        if not events_data and not isinstance(periods_payload, dict):
            return bad_request('Không có sự kiện nào để tạo')
        
        # Kiểm tra quyền cho vi phạm đột xuất (period "violation_sudden" hoặc "bonus_sudden")
        # Chỉ admin và dorm_supervisor mới được tạo/chỉnh sửa vi phạm đột xuất và điểm cộng đột xuất
        has_sudden_violations = False
        if isinstance(periods_payload, dict):
            has_sudden_violations = ('violation_sudden' in periods_payload and periods_payload['violation_sudden']) or \
                                   ('bonus_sudden' in periods_payload and periods_payload['bonus_sudden']) or \
                                   ('sudden' in periods_payload and periods_payload['sudden'])  # Backward compatibility
        elif events_data:
            # Kiểm tra trong events_data nếu có period là "violation_sudden", "bonus_sudden", "sudden" hoặc null
            for event_data in events_data:
                period = event_data.get('period')
                if period in ['violation_sudden', 'bonus_sudden', 'sudden'] or period is None:
                    has_sudden_violations = True
                    break
        
        if has_sudden_violations and user.role not in ['admin', 'dorm_supervisor']:
            return Response(
                {'error': 'Bạn không có quyền tạo/chỉnh sửa vi phạm đột xuất. Chỉ quản sinh mới có quyền này.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Lấy thông tin học sinh (không cần kiểm tra quyền)
        student_classroom_id = None
        if user.role == 'student':
            # Lấy classroom_id trực tiếp từ user document trong collection users
            users_coll = get_mongo_collection('users')
            try:
                user_doc = users_coll.find_one({'_id': ObjectId(user.id)})
            except:
                user_doc = users_coll.find_one({'id': str(user.id)})
            
            if not user_doc:
                return not_found('Không tìm thấy thông tin học sinh')
            
            student_classroom_id = user_doc.get('classroom_id')
            if not student_classroom_id:
                return bad_request('Học sinh chưa được phân lớp')
        
        # Tối ưu hóa: Nhóm events theo ngày và lớp để lưu trữ hiệu quả
        events_by_date_class = {}
        
        # Helper: push event into map
        def push_event(day_key: str, period_val, evt: dict):
            period_key = str(period_val)
            if day_key not in events_by_date_class:
                events_by_date_class[day_key] = {
                    'date': evt.get('date'),
                    'classroom_id': evt.get('classroom') or evt.get('classroom_id'),
                    'periods': {}
                }
            if period_key not in events_by_date_class[day_key]['periods']:
                events_by_date_class[day_key]['periods'][period_key] = []
            event_obj = {
                'event_type_key': evt.get('event_type_key'),
                'student_id': evt.get('student_id') or evt.get('student'),
                'points': evt.get('points', 0),
                'description': evt.get('description', ''),
            }
            # Add session field if present (for attendance: morning/afternoon)
            if evt.get('session'):
                event_obj['session'] = evt.get('session')
            events_by_date_class[day_key]['periods'][period_key].append(event_obj)

        # CASE A: periods payload provided (preferred)
        if isinstance(periods_payload, dict):
            # Determine date/classroom from top-level fields
            date = request.data.get('date')
            classroom_id = request.data.get('classroom') or request.data.get('classroom_id')
            if user.role == 'student':
                classroom_id = student_classroom_id
            if not date or not classroom_id:
                return bad_request('Ngày và lớp học là bắt buộc khi gửi periods')

            for period_key, events in periods_payload.items():
                if not isinstance(events, list):
                    continue
                
                # Đảm bảo day_key được tạo ngay cả khi events rỗng (để xử lý xóa period)
                day_key = f"{date}_{classroom_id}"
                if day_key not in events_by_date_class:
                    events_by_date_class[day_key] = {
                        'date': date,
                        'classroom_id': classroom_id,
                        'periods': {}
                    }
                
                # Nếu events rỗng, vẫn cần ghi nhận period này để xóa nó sau
                if len(events) == 0:
                    events_by_date_class[day_key]['periods'][period_key] = []
                else:
                    # Xử lý từng event
                    for ev in events:
                        # Map key->id if needed
                        et_id = ev.get('event_type')
                        et_key = ev.get('event_type_key')
                        # Xử lý custom_bonus_point: không cần tìm trong event_types collection
                        if not et_id and et_key and et_key != 'custom_bonus_point':
                            et_coll = get_mongo_collection('event_types')
                            et_doc = et_coll.find_one({'key': et_key})
                            if et_doc:
                                et_id = str(et_doc.get('_id'))
                                ev['event_type'] = et_id
                                if 'points' not in ev or ev.get('points') is None:
                                    ev['points'] = et_doc.get('default_points', 0)
                        # Compose day key and push
                        ev_comp = {
                            'event_type': et_id,  # Có thể là None cho custom_bonus_point
                            'event_type_key': et_key,
                            'student': ev.get('student_id') or ev.get('student'),
                            'student_id': ev.get('student_id') or ev.get('student'),  # Ensure student_id is present
                            'points': ev.get('points', 0),
                            'description': ev.get('description', ''),
                            'date': date,
                            'classroom': classroom_id,
                            'session': ev.get('session'),  # Support session field for attendance (morning/afternoon)
                        }
                        push_event(day_key, period_key, ev_comp)
        else:
            # CASE B: legacy flat events array
            for event_data in events_data:
                # Hỗ trợ nhận event_type_key và map sang id nếu thiếu
                et_id = event_data.get('event_type')
                et_key = event_data.get('event_type_key')
                # Xử lý custom_bonus_point: không cần tìm trong event_types collection
                if not et_id and et_key and et_key != 'custom_bonus_point':
                    et_coll = get_mongo_collection('event_types')
                    et_doc = et_coll.find_one({'key': et_key})
                    if et_doc:
                        et_id = str(et_doc.get('_id'))
                        event_data['event_type'] = et_id
                        if 'points' not in event_data or event_data.get('points') is None:
                            event_data['points'] = et_doc.get('default_points', 0)
                if not event_data.get('event_type') and not et_key:
                    continue

                date = event_data.get('date')
                classroom_id = student_classroom_id if user.role == 'student' else event_data.get('classroom')
                period = event_data.get('period')
                if not date or not classroom_id or not period:
                    continue
                day_key = f"{date}_{classroom_id}"
                ev_comp = {
                    'event_type': event_data.get('event_type'),
                    'event_type_key': et_key,
                    'student': event_data.get('student_id') or event_data.get('student'),
                    'points': event_data.get('points', 0),
                    'description': event_data.get('description', ''),
                    'date': date,
                    'classroom': classroom_id,
                }
                push_event(day_key, period, ev_comp)
            
        
        # Lưu trữ vào MongoDB với cấu trúc tối ưu
        events_coll = get_mongo_collection('events')
        created_events = []
        
        for key, day_data in events_by_date_class.items():
            # Tạo document cho mỗi ngày-lớp
            total_events = sum(len(period_events) for period_events in day_data['periods'].values())
            
            # Kiểm tra xem có điểm cộng đột xuất hoặc vi phạm đột xuất không
            has_sudden_events = 'bonus_sudden' in day_data['periods'] or 'violation_sudden' in day_data['periods'] or 'sudden' in day_data['periods']
            
            # Logic duyệt:
            # - Điểm cộng đột xuất và vi phạm đột xuất: luôn tự động duyệt (không cần duyệt)
            # - Hoạt động trong ngày: Học sinh cần duyệt, Giáo viên/Admin tự động duyệt
            if has_sudden_events:
                # Điểm cộng đột xuất và vi phạm đột xuất: tự động duyệt
                approval_status = 'approved'
                approved_by = str(user.id)
                approved_by_name = user.full_name or f"{user.first_name} {user.last_name}".strip()
            else:
                # Hoạt động trong ngày: học sinh cần duyệt, giáo viên/admin tự động duyệt
                approval_status = 'approved' if user.role in ['teacher', 'admin'] else 'pending'
                approved_by = str(user.id) if user.role in ['teacher', 'admin'] else None
                approved_by_name = user.full_name or f"{user.first_name} {user.last_name}".strip() if user.role in ['teacher', 'admin'] else None
            
            day_doc = {
                'date': day_data['date'],
                'classroom_id': day_data['classroom_id'],
                'periods': day_data['periods'],
                'total_events': total_events,
                'created_by': str(user.id),
                'created_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                'approval_status': approval_status,
                'approved_by': approved_by,
                'approved_by_name': approved_by_name,
                'approved_at': datetime.now().isoformat() if user.role in ['teacher', 'admin'] else None,
            }
            
            # Kiểm tra xem đã có document cho ngày-lớp này chưa
            existing = events_coll.find_one({
                'date': day_data['date'],
                'classroom_id': day_data['classroom_id']
            })
            
            if existing:
                # Merge periods thay vì replace để không mất dữ liệu cũ
                existing_periods = existing.get('periods', {})
                merged_periods = existing_periods.copy()
                
                # Merge từng period mới vào periods hiện có
                for period_key, new_events in day_data['periods'].items():
                    if period_key not in merged_periods:
                        # Period mới, thêm trực tiếp (chỉ nếu có events)
                        if new_events:
                            merged_periods[period_key] = new_events
                    else:
                        # Period đã tồn tại
                        if period_key == 'attendance':
                            # Với attendance, replace toàn bộ để đảm bảo xóa được events đã bị xóa ở frontend
                            if new_events:
                                merged_periods[period_key] = new_events
                            else:
                                # Nếu mảng rỗng, xóa period này
                                merged_periods.pop(period_key, None)
                        else:
                            # Với các period khác, merge events
                            existing_events = merged_periods[period_key]
                            # Tạo map để dễ tìm và replace events trùng lặp
                            # Key: (student_id, session) hoặc (student_id,) nếu không có session
                            event_map = {}
                            for ev in existing_events:
                                student_id = str(ev.get('student_id') or ev.get('student') or '')
                                session = ev.get('session', '')
                                key = (student_id, session) if session else (student_id,)
                                event_map[key] = ev
                            
                            # Thêm/thay thế events mới
                            for new_ev in new_events:
                                student_id = str(new_ev.get('student_id') or new_ev.get('student') or '')
                                session = new_ev.get('session', '')
                                key = (student_id, session) if session else (student_id,)
                                event_map[key] = new_ev
                            
                            # Chuyển lại thành list
                            merged_periods[period_key] = list(event_map.values())
                
                # Tính lại total_events sau khi merge
                total_events = sum(len(period_events) for period_events in merged_periods.values())
                
                # Cập nhật document hiện có
                update_data = {
                    'periods': merged_periods,
                    'total_events': total_events,
                    'updated_at': datetime.now().isoformat(),
                }
                
                # Cập nhật approval status nếu cần
                if user.role in ['teacher', 'admin']:
                    update_data.update({
                        'approval_status': 'approved',
                        'approved_by': str(user.id),
                        'approved_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
                        'approved_at': datetime.now().isoformat(),
                    })
                elif user.role == 'student':
                    # Học sinh update → cần duyệt lại
                    update_data.update({
                        'approval_status': 'pending',
                        'approved_by': None,
                        'approved_by_name': None,
                        'approved_at': None,
                    })
                
                events_coll.update_one(
                    {'_id': existing['_id']},
                    {'$set': update_data}
                )
                day_doc['_id'] = existing['_id']
                day_doc['periods'] = merged_periods
                day_doc['total_events'] = total_events
            else:
                # Tạo document mới
                result = events_coll.insert_one(day_doc)
                day_doc['_id'] = result.inserted_id
            
            created_events.append(to_plain(day_doc))
        
        return created({
            'message': f'Đã tạo/cập nhật {len(created_events)} ngày sự kiện',
            'created_count': len(created_events),
            'events': created_events
        })
        
    except Exception as exc:
        logger.exception('mongo_events_optimized_create error')
        return server_error(exc)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def mongo_events_optimized_replace(request):
    """Thay thế tất cả events cho một ngày-lớp cụ thể"""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        from bson import ObjectId
        from datetime import datetime
        
        user = request.user
        classroom_id = request.data.get('classroom_id') or request.data.get('classroom')
        date = request.data.get('date')
        events_data = request.data.get('events', [])
        periods_payload = request.data.get('periods')
        
        if not classroom_id or not date:
            return bad_request('Thiếu classroom_id hoặc date')
        
        # Kiểm tra quyền cho vi phạm đột xuất (period "sudden")
        # Chỉ admin và dorm_supervisor mới được thay thế vi phạm đột xuất
        has_sudden_violations = False
        if isinstance(periods_payload, dict):
            has_sudden_violations = 'sudden' in periods_payload and periods_payload['sudden']
        elif events_data:
            for event_data in events_data:
                period = event_data.get('period')
                if period == 'sudden' or period is None:
                    has_sudden_violations = True
                    break
        
        if has_sudden_violations and user.role not in ['admin', 'dorm_supervisor']:
            return Response(
                {'error': 'Bạn không có quyền thay thế vi phạm đột xuất. Chỉ quản sinh mới có quyền này.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Tự động lấy lớp của học sinh nếu là student
        if user.role == 'student':
            students_coll = get_mongo_collection('students')
            student_doc = students_coll.find_one({'user.id': str(user.id)})
            
            if not student_doc:
                return not_found('Không tìm thấy thông tin học sinh')
            
            student_classroom_id = student_doc.get('classroom', {}).get('id')
            if not student_classroom_id:
                return bad_request('Học sinh chưa được phân lớp')
            
            # Học sinh chỉ có thể thay thế events của lớp mình
            if classroom_id != student_classroom_id:
                return Response({'error': 'Bạn chỉ có thể thay thế events của lớp mình'}, status=status.HTTP_403_FORBIDDEN)
        
        # Build periods structure (accept both 'periods' and 'events')
        periods = {}
        et_coll = get_mongo_collection('event_types')
        if isinstance(periods_payload, dict):
            for period_key, events in periods_payload.items():
                pk = str(period_key)
                if pk not in periods:
                    periods[pk] = []
                if not isinstance(events, list):
                    continue
                for ev in events:
                    et_id = ev.get('event_type')
                    et_key = ev.get('event_type_key')
                    # Xử lý custom_bonus_point: không cần tìm trong event_types collection
                    if not et_id and et_key and et_key != 'custom_bonus_point':
                        et_doc = et_coll.find_one({'key': et_key})
                        if et_doc:
                            et_id = str(et_doc.get('_id'))
                            ev['event_type'] = et_id
                            if 'points' not in ev or ev.get('points') is None:
                                ev['points'] = et_doc.get('default_points', 0)
            periods[pk].append({
                        'event_type_key': et_key,
                        'student_id': ev.get('student'),
                        'points': ev.get('points', 0),
                        'description': ev.get('description', ''),
                    })
        else:
            for event_data in events_data:
                period = str(event_data.get('period'))
                if period not in periods:
                    periods[period] = []
                et_id = event_data.get('event_type')
                et_key = event_data.get('event_type_key')
                # Xử lý custom_bonus_point: không cần tìm trong event_types collection
                if not et_id and et_key and et_key != 'custom_bonus_point':
                    et_doc = et_coll.find_one({'key': et_key})
                    if et_doc:
                        et_id = str(et_doc.get('_id'))
                        event_data['event_type'] = et_id
                        if 'points' not in event_data or event_data.get('points') is None:
                            event_data['points'] = et_doc.get('default_points', 0)
                if et_id and not et_key:
                    try:
                        from bson import ObjectId
                        et_doc2 = et_coll.find_one({'_id': ObjectId(et_id)})
                        if et_doc2:
                            et_key = et_doc2.get('key')
                    except Exception:
                        pass
                event_obj = {
                    'event_type_key': et_key,
                    'student_id': event_data.get('student'),
                    'points': event_data.get('points', 0),
                    'description': event_data.get('description', ''),
                }
                # Add session field if present (for attendance: morning/afternoon)
                if event_data.get('session'):
                    event_obj['session'] = event_data.get('session')
                periods[period].append(event_obj)
        
        # Tạo document mới
        day_doc = {
            'date': date,
            'classroom_id': classroom_id,
            'periods': periods,
            'total_events': len(events_data),
            'created_by': str(user.id),
            'created_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
        }
        
        # Thay thế document
        events_coll = get_mongo_collection('events')
        events_coll.replace_one(
            {'date': date, 'classroom_id': classroom_id},
            day_doc,
            upsert=True
        )
        
        return ok({
            'message': f'Đã thay thế {len(events_data)} events cho ngày {date}',
            'events': to_plain(day_doc)
        })
        
    except Exception as exc:
        logger.exception('mongo_events_optimized_replace error')
        return server_error(exc)

# =============================================================================
# BULK OPERATIONS - MongoDB
# =============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mongo_events_bulk_sync(request):
    """Đồng bộ events cho một period cụ thể"""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        from bson import ObjectId
        from datetime import datetime
        
        user = request.user
        classroom_id = request.data.get('classroom_id')
        date = request.data.get('date')
        period = request.data.get('period')
        events_data = request.data.get('events', [])
        
        if not classroom_id or not date or not period:
            return Response({'error': 'Thiếu thông tin bắt buộc'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Kiểm tra quyền cho vi phạm đột xuất (period "sudden")
        # Chỉ admin và dorm_supervisor mới được đồng bộ vi phạm đột xuất
        if period == 'sudden' and user.role not in ['admin', 'dorm_supervisor']:
            return Response(
                {'error': 'Bạn không có quyền đồng bộ vi phạm đột xuất. Chỉ quản sinh mới có quyền này.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Tự động lấy lớp của học sinh nếu là student
        if user.role == 'student':
            students_coll = get_mongo_collection('students')
            student_doc = students_coll.find_one({'user.id': str(user.id)})
            
            if not student_doc:
                return Response({'error': 'Không tìm thấy thông tin học sinh'}, status=status.HTTP_404_NOT_FOUND)
            
            student_classroom_id = student_doc.get('classroom', {}).get('id')
            if not student_classroom_id:
                return Response({'error': 'Học sinh chưa được phân lớp'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Học sinh chỉ có thể sync events của lớp mình
            if classroom_id != student_classroom_id:
                return Response({'error': 'Bạn chỉ có thể sync events của lớp mình'}, status=status.HTTP_403_FORBIDDEN)
        
        # Tìm document hiện có
        events_coll = get_mongo_collection('events')
        existing_doc = events_coll.find_one({
            'date': date,
            'classroom_id': classroom_id
        })
        
        if existing_doc:
            # Cập nhật period cụ thể
            events_coll.update_one(
                {'_id': existing_doc['_id']},
                {
                    '$set': {
                        f'periods.{period}': events_data,
                        'updated_at': datetime.now().isoformat(),
                    }
                }
            )
        else:
            # Tạo document mới
            day_doc = {
                'date': date,
                'classroom_id': classroom_id,
                'periods': {str(period): events_data},
                'total_events': len(events_data),
                'created_by': str(user.id),
                'created_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
            }
            events_coll.insert_one(day_doc)
        
        return Response({
            'message': f'Đã đồng bộ {len(events_data)} events cho tiết {period}',
            'period': period,
            'events_count': len(events_data)
        })
        
    except Exception as exc:
        logger.exception('mongo_events_bulk_sync error')
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mongo_events_bulk_replace(request):
    """Ghi đè toàn bộ events cho một ngày (tất cả 7 tiết)"""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        from bson import ObjectId
        from datetime import datetime
        
        user = request.user
        classroom_id = request.data.get('classroom_id')
        date = request.data.get('date')
        periods_data = request.data.get('periods', {})
        
        if not classroom_id or not date:
            return Response({'error': 'Thiếu thông tin bắt buộc'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Kiểm tra quyền cho vi phạm đột xuất (period "violation_sudden" hoặc "bonus_sudden")
        # Chỉ admin và dorm_supervisor mới được ghi đè vi phạm đột xuất và điểm cộng đột xuất
        has_sudden_violations = ('violation_sudden' in periods_data and periods_data['violation_sudden']) or \
                               ('bonus_sudden' in periods_data and periods_data['bonus_sudden']) or \
                               ('sudden' in periods_data and periods_data['sudden'])  # Backward compatibility
        if has_sudden_violations and user.role not in ['admin', 'dorm_supervisor']:
            return Response(
                {'error': 'Bạn không có quyền ghi đè vi phạm đột xuất. Chỉ quản sinh mới có quyền này.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Tự động lấy lớp của học sinh nếu là student
        if user.role == 'student':
            students_coll = get_mongo_collection('students')
            student_doc = students_coll.find_one({'user.id': str(user.id)})
            
            if not student_doc:
                return Response({'error': 'Không tìm thấy thông tin học sinh'}, status=status.HTTP_404_NOT_FOUND)
            
            student_classroom_id = student_doc.get('classroom', {}).get('id')
            if not student_classroom_id:
                return Response({'error': 'Học sinh chưa được phân lớp'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Học sinh chỉ có thể replace events của lớp mình
            if classroom_id != student_classroom_id:
                return Response({'error': 'Bạn chỉ có thể replace events của lớp mình'}, status=status.HTTP_403_FORBIDDEN)
        
        # Tìm document hiện có
        events_coll = get_mongo_collection('events')
        existing_doc = events_coll.find_one({
            'date': date,
            'classroom_id': classroom_id
        })
        
        # Tính tổng số events (chỉ tính các period có events, không tính mảng rỗng)
        total_events = sum(len(events) for events in periods_data.values() if isinstance(events, list) and len(events) > 0)
        
        # Kiểm tra xem có điểm cộng đột xuất hoặc vi phạm đột xuất không (chỉ tính các period có events)
        has_sudden_events = any(
            (key in periods_data and isinstance(periods_data[key], list) and len(periods_data[key]) > 0)
            for key in ['bonus_sudden', 'violation_sudden', 'sudden']
        )
        
        # Logic duyệt:
        # - Điểm cộng đột xuất và vi phạm đột xuất: luôn tự động duyệt (không cần duyệt)
        # - Hoạt động trong ngày: Học sinh cần duyệt, Giáo viên/Admin tự động duyệt
        if has_sudden_events:
            # Điểm cộng đột xuất và vi phạm đột xuất: tự động duyệt
            approval_status = 'approved'
            approved_by = str(user.id)
            approved_by_name = user.full_name or f"{user.first_name} {user.last_name}".strip()
        else:
            # Hoạt động trong ngày: học sinh cần duyệt, giáo viên/admin tự động duyệt
            approval_status = 'approved' if user.role in ['teacher', 'admin'] else 'pending'
            approved_by = str(user.id) if user.role in ['teacher', 'admin'] else None
            approved_by_name = user.full_name or f"{user.first_name} {user.last_name}".strip() if user.role in ['teacher', 'admin'] else None
        
        if existing_doc:
            # Ghi đè toàn bộ periods, nhưng loại bỏ các period có mảng rỗng
            # Lọc ra các period có mảng rỗng để xóa chúng
            periods_to_set = {}
            periods_to_unset = []
            
            for period_key, period_events in periods_data.items():
                if isinstance(period_events, list) and len(period_events) == 0:
                    # Mảng rỗng: xóa period này
                    periods_to_unset.append(f'periods.{period_key}')
                else:
                    # Có dữ liệu: giữ lại
                    periods_to_set[period_key] = period_events
            
            update_data = {
                'total_events': total_events,
                'updated_at': datetime.now().isoformat(),
            }
            
            # Luôn set periods (có thể là dict rỗng nếu tất cả periods đều bị xóa)
            # Nếu có periods để set, dùng periods_to_set; nếu không, dùng dict rỗng để xóa tất cả periods
            if periods_to_set:
                update_data['periods'] = periods_to_set
            elif periods_to_unset:
                # Nếu chỉ có periods để unset và không có periods để set, set periods thành dict rỗng
                # MongoDB sẽ xóa các field được unset và set periods thành {}
                update_data['periods'] = {}
            
            # Cập nhật approval status
            if has_sudden_events:
                # Điểm cộng đột xuất và vi phạm đột xuất: tự động duyệt
                update_data.update({
                    'approval_status': 'approved',
                    'approved_by': str(user.id),
                    'approved_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
                    'approved_at': datetime.now().isoformat(),
                })
            elif user.role in ['teacher', 'admin']:
                # Giáo viên/Admin: tự động duyệt
                update_data.update({
                    'approval_status': 'approved',
                    'approved_by': str(user.id),
                    'approved_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
                    'approved_at': datetime.now().isoformat(),
                })
            elif user.role == 'student':
                # Học sinh update hoạt động trong ngày → cần duyệt lại
                update_data.update({
                    'approval_status': 'pending',
                    'approved_by': None,
                    'approved_by_name': None,
                    'approved_at': None,
                })
            
            # Build update operation
            update_operation = {}
            if update_data:
                update_operation['$set'] = update_data
            if periods_to_unset:
                update_operation['$unset'] = {period: '' for period in periods_to_unset}
            
            events_coll.update_one(
                {'_id': existing_doc['_id']},
                update_operation
            )
            action = 'updated'
        else:
            # Tạo document mới
            day_doc = {
                'date': date,
                'classroom_id': classroom_id,
                'periods': periods_data,
                'total_events': total_events,
                'created_by': str(user.id),
                'created_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                'approval_status': approval_status,
                'approved_by': approved_by,
                'approved_by_name': approved_by_name,
                'approved_at': datetime.now().isoformat() if user.role in ['teacher', 'admin'] else None,
            }
            events_coll.insert_one(day_doc)
            action = 'created'
        
        return Response({
            'message': f'Đã {action} {total_events} events cho {len(periods_data)} tiết',
            'total_events': total_events,
            'periods_count': len(periods_data),
            'action': action
        })
        
    except Exception as exc:
        logger.exception('mongo_events_bulk_replace error')
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mongo_events_approve(request):
    """Duyệt events - chỉ admin và giáo viên chủ nhiệm mới có quyền"""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        from bson import ObjectId
        from datetime import datetime
        
        user = request.user
        
        # Chỉ admin và giáo viên mới có quyền duyệt
        if user.role not in ['admin', 'teacher']:
            return Response({'error': 'Bạn không có quyền duyệt sự kiện'}, status=status.HTTP_403_FORBIDDEN)
        
        event_id = request.data.get('event_id')
        action = request.data.get('action')  # 'approve' hoặc 'reject'
        
        if not event_id or not action:
            return Response({'error': 'Thiếu thông tin bắt buộc'}, status=status.HTTP_400_BAD_REQUEST)
        
        if action not in ['approve', 'reject']:
            return Response({'error': 'Action không hợp lệ'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Tìm event document
        events_coll = get_mongo_collection('events')
        try:
            event_doc = events_coll.find_one({'_id': ObjectId(event_id)})
        except:
            return Response({'error': 'ID không hợp lệ'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not event_doc:
            return Response({'error': 'Không tìm thấy sự kiện'}, status=status.HTTP_404_NOT_FOUND)
        
        # Kiểm tra quyền duyệt
        if user.role == 'teacher':
            # Giáo viên chỉ có thể duyệt events của lớp mình chủ nhiệm
            classrooms_coll = get_mongo_collection('classrooms')
            teacher_classrooms = classrooms_coll.find({'homeroom_teacher.id': str(user.id)})
            teacher_classroom_ids = [str(doc['_id']) for doc in teacher_classrooms]
            
            if event_doc.get('classroom_id') not in teacher_classroom_ids:
                return Response({'error': 'Bạn chỉ có thể duyệt sự kiện của lớp mình chủ nhiệm'}, status=status.HTTP_403_FORBIDDEN)
        
        # Cập nhật approval status
        update_data = {
            'approval_status': 'approved' if action == 'approve' else 'rejected',
            'approved_by': str(user.id),
            'approved_by_name': user.full_name or f"{user.first_name} {user.last_name}".strip(),
            'approved_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
        }
        
        events_coll.update_one(
            {'_id': ObjectId(event_id)},
            {'$set': update_data}
        )
        
        return Response({
            'message': f'Đã {action} sự kiện thành công',
            'event_id': event_id,
            'approval_status': update_data['approval_status'],
            'approved_by': update_data['approved_by_name'],
            'approved_at': update_data['approved_at']
        })
        
    except Exception as exc:
        logger.exception('mongo_events_approve error')
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])  # Public API - không cần authentication
def mongo_events_public(request):
    """Public API để xem sự kiện toàn trường - không cần authentication"""
    try:
        from applications.common.mongo import get_mongo_collection, to_plain
        from bson import ObjectId
        from datetime import datetime
        
        # Lấy parameters
        date = request.GET.get('date')
        classroom_id = request.GET.get('classroom_id')
        
        if not date:
            return Response({'error': 'Thiếu tham số date'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Query events
        events_coll = get_mongo_collection('events')
        classrooms_coll = get_mongo_collection('classrooms')
        users_coll = get_mongo_collection('users')
        
        # Build query
        query = {
            'date': date
            # Removed approval_status filter to see all events first
        }
        
        if classroom_id and classroom_id != 'all':
            query['classroom_id'] = classroom_id
        
        print(f"Public API - Final query: {query}")
        print(f"Public API - Date filter: {date}")
        print(f"Public API - Classroom filter: {classroom_id}")
        
        # Pagination parameters
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        page_size = min(page_size, 100)  # Limit max page size to 100
        
        print(f"Public API - Pagination params: page={page}, page_size={page_size}")
        print(f"Public API - Query: {query}")
        
        # Count total documents
        total_count = events_coll.count_documents(query)
        total_pages = (total_count + page_size - 1) // page_size
        
        print(f"Public API - Total count: {total_count}, Total pages: {total_pages}")
        
        # Calculate skip
        skip = (page - 1) * page_size
        
        print(f"Public API - Skip: {skip}, Limit: {page_size}")
        
        # Get paginated results
        cursor = events_coll.find(query).sort('created_at', -1).skip(skip).limit(page_size)
        event_docs = list(cursor)
        
        print(f"Public API - Retrieved {len(event_docs)} documents")
        print(f"Public API - First few doc IDs: {[str(doc.get('_id', 'no_id')) for doc in event_docs[:3]]}")
        print(f"Public API - Query executed: {query}")
        print(f"Public API - Skip: {skip}, Limit: {page_size}")
        
        # Debug: Check if we're getting more than expected
        if len(event_docs) > page_size:
            print(f"Public API - WARNING: Got {len(event_docs)} docs but limit was {page_size}")
        
        # Process events
        processed_events = []
        event_counter = 0  # Counter for unique IDs
        for event_doc in event_docs:
            event_plain = to_plain(event_doc)
            
            # Lấy thông tin classroom
            classroom_info = None
            if event_plain.get('classroom_id'):
                classroom_doc = classrooms_coll.find_one({'_id': ObjectId(event_plain['classroom_id'])})
                if classroom_doc:
                    classroom_info = {
                        'id': str(classroom_doc['_id']),
                        'name': classroom_doc.get('name', ''),
                        'full_name': classroom_doc.get('full_name', ''),
                        'grade': classroom_doc.get('grade', '')
                    }
            
            # Process periods để tạo individual events
            periods = event_plain.get('periods', {})
            for period_num, period_events in periods.items():
                if not isinstance(period_events, list):
                    continue
                for event in period_events:
                    # Lấy thông tin student nếu có
                    student_info = None
                    if event.get('student_id'):
                        student_doc = users_coll.find_one({
                            '_id': ObjectId(event['student_id']),
                            'role': 'student'
                        })
                        if student_doc:
                            student_info = {
                                'id': str(student_doc['_id']),
                                'full_name': student_doc.get('full_name', '')
                            }
                    
                    # Tạo event object với unique ID
                    event_id = event.get('id', f"event_{event_counter}")
                    processed_event = {
                        'id': f"{event_plain['id']}_{period_num}_{event_id}_{event_counter}",
                        'title': event.get('title', ''),
                        'description': event.get('description', ''),
                        'date': event_plain['date'],
                        'classroom': classroom_info or {
                            'id': event_plain.get('classroom_id', ''),
                            'name': 'Unknown',
                            'full_name': 'Unknown',
                            'grade': ''
                        },
                        'student': student_info,
                        'event_type': event.get('event_type', ''),
                        'points': event.get('points', 0),
                        'created_at': event_plain.get('created_at', ''),
                        'period': int(period_num),
                        'approval_status': event_plain.get('approval_status', 'approved'),
                        'approved_by': event_plain.get('approved_by'),
                        'approved_by_name': event_plain.get('approved_by_name'),
                        'approved_at': event_plain.get('approved_at')
                    }
                    
                    processed_events.append(processed_event)
                    event_counter += 1  # Increment counter for next event
        
        # Build pagination URLs
        base_url = request.build_absolute_uri().split('?')[0]
        params = request.GET.copy()
        
        next_url = None
        if page < total_pages:
            params['page'] = page + 1
            next_url = f"{base_url}?{params.urlencode()}"
        
        previous_url = None
        if page > 1:
            params['page'] = page - 1
            previous_url = f"{base_url}?{params.urlencode()}"
        
        return Response({
            'events': processed_events,
            'total': len(processed_events),
            'total_count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'next': next_url,
            'previous': previous_url,
            'date': date,
            'classroom_id': classroom_id
        })
        
    except Exception as exc:
        logger.exception('mongo_events_public error')
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def derive_attendance_type(morning: str, afternoon: str) -> str:
    """
    Suy luận loại nghỉ từ morning và afternoon.
    Returns: attendance code (legacy format hoặc session-based)
    """
    has_morning = morning and morning in ['attendance_sp', 'attendance_sk']
    has_afternoon = afternoon and afternoon in ['attendance_cp', 'attendance_ck']
    
    if has_morning and has_afternoon:
        # Nghỉ cả ngày
        if morning == 'attendance_sp' and afternoon == 'attendance_cp':
            return 'attendance_spcp'
        elif morning == 'attendance_sk' and afternoon == 'attendance_ck':
            return 'attendance_skck'
        elif morning == 'attendance_sp' and afternoon == 'attendance_ck':
            return 'attendance_spck'
        elif morning == 'attendance_sk' and afternoon == 'attendance_cp':
            return 'attendance_skcp'
    elif has_morning:
        return morning  # SP hoặc SK
    elif has_afternoon:
        return afternoon  # CP hoặc CK
    
    return ''

def get_absence_periods_from_sessions(morning: str, afternoon: str) -> int:
    """Tính số buổi nghỉ từ morning và afternoon"""
    periods = 0
    if morning and morning in ['attendance_sp', 'attendance_sk']:
        periods += 1
    if afternoon and afternoon in ['attendance_cp', 'attendance_ck']:
        periods += 1
    return periods

def get_excused_unexcused_periods(morning: str, afternoon: str) -> dict:
    """Tính số buổi có phép và không phép"""
    excused = 0
    unexcused = 0
    
    if morning == 'attendance_sp':
        excused += 1
    elif morning == 'attendance_sk':
        unexcused += 1
    
    if afternoon == 'attendance_cp':
        excused += 1
    elif afternoon == 'attendance_ck':
        unexcused += 1
    
    return {'excused': excused, 'unexcused': unexcused}

def get_absence_periods(attendance_code: str) -> int:
    """
    Calculate number of absence periods (sessions) from attendance code.
    1 day = 2 periods (morning + afternoon)
    Full day absence = 2 periods, Half day absence = 1 period
    """
    # Full day absence: 2 periods
    # spck = sáng có phép + chiều không phép = 2 buổi
    # skcp = sáng không phép + chiều có phép = 2 buổi
    if attendance_code in ['attendance_spcp', 'attendance_skck', 
                          'attendance_spck', 'attendance_skcp']:
        return 2
    # Half day absence: 1 period
    if attendance_code in ['attendance_sp', 'attendance_cp', 'attendance_sk', 'attendance_ck']:
        return 1
    return 0

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mongo_attendance_export(request):
    """Xuất điểm danh ra file Excel theo template"""
    try:
        from django.http import HttpResponse
        from calendar import monthrange
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        classroom_id = request.query_params.get('classroom_id')
        month = int(request.query_params.get('month', datetime.now().month))
        year = int(request.query_params.get('year', datetime.now().year))
        
        if not classroom_id:
            return bad_request('classroom_id là bắt buộc')
        
        # Get classroom info
        classrooms_coll = get_mongo_collection('classrooms')
        classroom = classrooms_coll.find_one({'_id': ObjectId(classroom_id)})
        if not classroom:
            return not_found('Không tìm thấy lớp học')
        
        # Get students for this class
        users_coll = get_mongo_collection('users')
        students = list(users_coll.find({
            'role': 'student',
            'classroom_id': classroom_id
        }).sort('full_name', 1))
        
        # Get attendance data for the month
        events_coll = get_mongo_collection('events')
        start_date = f"{year}-{month:02d}-01"
        days_in_month = monthrange(year, month)[1]
        end_date = f"{year}-{month:02d}-{days_in_month:02d}"
        
        attendance_data = {}
        events = events_coll.find({
            'classroom_id': classroom_id,
            'date': {'$gte': start_date, '$lte': end_date},
            'periods.attendance': {'$exists': True, '$ne': []}
        })
        
        for event_doc in events:
            date = event_doc.get('date')
            attendance_events = event_doc.get('periods', {}).get('attendance', [])
            if date not in attendance_data:
                attendance_data[date] = {}
            
            # Group attendance events by student_id to handle session-based data
            student_sessions = {}
            for att_event in attendance_events:
                student_id = att_event.get('student_id')
                # Convert ObjectId to string if needed
                if student_id and hasattr(student_id, '__str__'):
                    student_id = str(student_id)
                else:
                    student_id = str(student_id) if student_id else ''
                
                if not student_id:
                    continue
                
                event_type_key = att_event.get('event_type_key', '')
                session = att_event.get('session', 'morning')  # Default to morning if not specified
                
                if student_id not in student_sessions:
                    student_sessions[student_id] = {'morning': '', 'afternoon': ''}
                
                # Map to morning/afternoon based on session or event type
                if session == 'morning' or event_type_key in ['attendance_sp', 'attendance_sk']:
                    student_sessions[student_id]['morning'] = event_type_key
                elif session == 'afternoon' or event_type_key in ['attendance_cp', 'attendance_ck']:
                    student_sessions[student_id]['afternoon'] = event_type_key
                else:
                    # Legacy format - try to derive from event_type_key
                    legacy_map = {
                        'attendance_spcp': {'morning': 'attendance_sp', 'afternoon': 'attendance_cp'},
                        'attendance_skck': {'morning': 'attendance_sk', 'afternoon': 'attendance_ck'},
                        'attendance_spck': {'morning': 'attendance_sp', 'afternoon': 'attendance_ck'},
                        'attendance_skcp': {'morning': 'attendance_sk', 'afternoon': 'attendance_cp'},
                    }
                    if event_type_key in legacy_map:
                        mapped = legacy_map[event_type_key]
                        student_sessions[student_id]['morning'] = mapped['morning']
                        student_sessions[student_id]['afternoon'] = mapped['afternoon']
                    else:
                        # Fallback: store as-is (legacy single code)
                        attendance_data[date][student_id] = event_type_key
            
            # Derive attendance type from sessions for each student
            for student_id, sessions in student_sessions.items():
                morning = sessions.get('morning', '')
                afternoon = sessions.get('afternoon', '')
                
                if morning or afternoon:
                    # Use helper function to derive attendance type
                    derived_type = derive_attendance_type(morning, afternoon)
                    if derived_type:
                        attendance_data[date][student_id] = derived_type
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Điểm danh"
        
        # Header styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header rows
        ws.merge_cells('A1:F1')
        ws['A1'] = 'SỞ GIÁO DỤC VÀ ĐÀO TẠO ĐỒNG THÁP'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        
        ws.merge_cells('G1:L1')
        ws['G1'] = 'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM'
        ws['G1'].font = header_font
        ws['G1'].alignment = center_align
        
        ws.merge_cells('A2:F2')
        ws['A2'] = 'TRƯỜNG THPT LAI VUNG 3'
        ws['A2'].font = header_font
        ws['A2'].alignment = center_align
        
        ws.merge_cells('G2:L2')
        ws['G2'] = 'Độc lập - Tự do - Hạnh phúc'
        ws['G2'].font = header_font
        ws['G2'].alignment = center_align
        
        ws.merge_cells('A3:L3')
        ws['A3'] = 'SỐ ĐIỂM DANH'
        ws['A3'].font = title_font
        ws['A3'].alignment = center_align
        
        # Month and class info
        month_names = ['', 'Tháng 1', 'Tháng 2', 'Tháng 3', 'Tháng 4', 'Tháng 5', 'Tháng 6',
                      'Tháng 7', 'Tháng 8', 'Tháng 9', 'Tháng 10', 'Tháng 11', 'Tháng 12']
        classroom_name = classroom.get('full_name', '')
        grade = classroom.get('grade', '')
        
        ws.merge_cells('A4:L4')
        ws['A4'] = f'{month_names[month]} năm {year} - Khối {grade} - {classroom_name}'
        ws['A4'].font = header_font
        ws['A4'].alignment = center_align
        
        # Get actual days in month
        days_in_month = monthrange(year, month)[1]
        
        # Day names in Vietnamese
        day_names = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
        
        # Column headers - Row 6 (main headers)
        row = 6
        col = 1
        
        # Fixed columns (merge 2 rows)
        # Set value first, then merge
        cell = ws.cell(row=row, column=col, value='STT')
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)
        col += 1
        
        cell = ws.cell(row=row, column=col, value='Mã học sinh')
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)
        col += 1
        
        cell = ws.cell(row=row, column=col, value='Họ và tên')
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)
        col += 1
        
        # Day columns (only actual days in month)
        # Row 6: Day numbers, Row 7: Day names
        day_cols = {}  # Track column for each day
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
                day_name = day_names[weekday]
            except:
                day_name = ''
            
            day_cols[day] = col
            # Row 6: Day number
            cell = ws.cell(row=row, column=col, value=str(day))
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            col += 1
        
        # Summary columns - TS Buổi nghỉ (merge 3 columns, not rows)
        summary_start_col = col
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
        cell = ws.cell(row=row, column=col, value='TS Buổi nghỉ')
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        col += 3
        
        # TS bỏ tiết (no merge)
        cell = ws.cell(row=row, column=col, value='TS bỏ tiết')
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        
        # Row 7 (sub-headers)
        row = 7
        col = 1
        
        # Fixed columns (already merged, so skip)
        col += 3
        
        # Day names (thứ) for each day
        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
                day_name = day_names[weekday]
            except:
                day_name = ''
            
            cell = ws.cell(row=row, column=col, value=day_name)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            col += 1
        
        # Summary sub-headers (TS, P, K)
        summary_sub_headers = ['TS', 'P', 'K']
        for sub_header in summary_sub_headers:
            cell = ws.cell(row=row, column=col, value=sub_header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            col += 1
        
        # TS bỏ tiết (already merged, so skip)
        
        # Write student data
        row = 8
        for idx, student in enumerate(students, 1):
            student_obj_id = student.get('_id')
            # Convert ObjectId to string
            if hasattr(student_obj_id, '__str__'):
                student_id = str(student_obj_id)
            else:
                student_id = str(student_obj_id) if student_obj_id else ''
            student_code = student.get('student_code', '')
            full_name = student.get('full_name', '')
            
            # Split name
            name_parts = full_name.split()
            last_name = ' '.join(name_parts[:-1]) if len(name_parts) > 1 else ''
            first_name = name_parts[-1] if name_parts else full_name
            
            ws.cell(row=row, column=1, value=idx).border = border  # STT
            ws.cell(row=row, column=2, value=student_code).border = border  # Mã học sinh
            ws.cell(row=row, column=3, value=full_name).border = border  # Họ và tên
            
            # Write attendance for each day (only actual days in month)
            col = 4
            total_absence_periods = 0  # Total absence periods (sessions), not days
            total_excused_periods = 0   # Total excused absence periods
            total_unexcused_periods = 0 # Total unexcused absence periods
            
            # Map attendance code to display
            code_map = {
                'attendance_spcp': 'spcp',
                'attendance_sp': 'sp',
                'attendance_cp': 'cp',
                'attendance_skck': 'skck',
                'attendance_sk': 'sk',
                'attendance_ck': 'ck',
                'attendance_spck': 'spck',
                'attendance_skcp': 'skcp'
            }
            
            for day in range(1, days_in_month + 1):
                date_str = f"{year}-{month:02d}-{day:02d}"
                attendance_code = attendance_data.get(date_str, {}).get(student_id, '')
                
                # Write attendance code if exists
                if attendance_code:
                    display_code = code_map.get(attendance_code, '')
                    ws.cell(row=row, column=col, value=display_code).border = border
                    if display_code:
                        # Calculate absence periods (sessions) for this day
                        periods = get_absence_periods(attendance_code)
                        total_absence_periods += periods
                        
                        # Count excused and unexcused periods
                        # For legacy codes, use direct mapping
                        # For session-based codes, derive from sessions if needed
                        if attendance_code in ['attendance_spcp', 'attendance_skck', 'attendance_spck', 'attendance_skcp']:
                            # Legacy full-day codes
                            if attendance_code == 'attendance_spcp':
                                total_excused_periods += 2
                            elif attendance_code == 'attendance_skck':
                                total_unexcused_periods += 2
                            elif attendance_code == 'attendance_spck':
                                # Sáng có phép + chiều không phép = 1 buổi có phép + 1 buổi không phép
                                total_excused_periods += 1
                                total_unexcused_periods += 1
                            elif attendance_code == 'attendance_skcp':
                                # Sáng không phép + chiều có phép = 1 buổi không phép + 1 buổi có phép
                                total_unexcused_periods += 1
                                total_excused_periods += 1
                        else:
                            # Session-based codes (SP, SK, CP, CK) - use helper function
                            # Since we already derived the type, we can use direct mapping
                            if attendance_code == 'attendance_sp':
                                total_excused_periods += 1
                            elif attendance_code == 'attendance_sk':
                                total_unexcused_periods += 1
                            elif attendance_code == 'attendance_cp':
                                total_excused_periods += 1
                            elif attendance_code == 'attendance_ck':
                                total_unexcused_periods += 1
                else:
                    # Empty cell
                    ws.cell(row=row, column=col, value='').border = border
                col += 1
            
            # Summary columns - TS Buổi nghỉ (total absence periods)
            ws.cell(row=row, column=col, value=total_absence_periods).border = border  # TS (total periods)
            ws.cell(row=row, column=col+1, value=total_excused_periods).border = border  # P (excused periods)
            ws.cell(row=row, column=col+2, value=total_unexcused_periods).border = border  # K (unexcused periods)
            ws.cell(row=row, column=col+3, value=0).border = border  # TS bỏ tiết
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5  # STT
        ws.column_dimensions['B'].width = 15  # Mã học sinh
        ws.column_dimensions['C'].width = 20  # Họ và tên
        ws.column_dimensions['D'].width = 15  # Tên
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'diem_danh_{classroom_name}_{month}_{year}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as exc:
        logger.exception('mongo_attendance_export error')
        return server_error(exc)
